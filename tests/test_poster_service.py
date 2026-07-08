from __future__ import annotations

from pathlib import Path
import tempfile
import unittest
from zipfile import ZipFile

from wp_auto_poster_gui.core.models import Post, PostResult, PosterOptions, UploadedMedia, WordPressConfig
from wp_auto_poster_gui.core.poster_service import export_links_to_source_excel
from wp_auto_poster_gui.core.poster_service import list_website_posts
from wp_auto_poster_gui.core.poster_service import publish_existing_posts_bulk
from wp_auto_poster_gui.core.poster_service import publish_posts
from wp_auto_poster_gui.core.poster_service import publish_from_excel
from wp_auto_poster_gui.core.poster_service import update_existing_posts_image_layout
from wp_auto_poster_gui.core.poster_service import publish_website_posts_bulk
from wp_auto_poster_gui.core.poster_service import update_website_posts_image_layout


class FakeClient:
    def __init__(self):
        self.created = []
        self.updated = []

    def find_post_by_title(self, title: str):
        if title == "Duplicate":
            return {"id": 99, "link": "https://example.com/duplicate"}
        return None

    def upload_media_from_path(self, path: Path) -> UploadedMedia:
        return UploadedMedia(100 + len(str(path)), f"https://cdn.example.com/{path.name}", path.name)

    def upload_media_from_url(self, url: str) -> UploadedMedia:
        return UploadedMedia(55, url, "remote.jpg")

    def create_post(self, post: Post, content: str, featured_media_id: int | None = None):
        self.created.append((post, content, featured_media_id))
        return {"link": f"https://example.com/{post.row_number}"}

    def update_post(self, post_id: int, post: Post, content: str, featured_media_id: int | None = None):
        self.updated.append((post_id, post, content, featured_media_id))
        return {"link": f"https://example.com/existing/{post_id}"}


class PosterServiceTest(unittest.TestCase):
    def test_publishes_with_local_images_and_updates_duplicate_seo(self) -> None:
        fake = FakeClient()
        posts = [
            Post(2, "First", "<p>A</p><p>B</p>", ma_bai="bai01"),
            Post(3, "Duplicate", "C", ma_bai="bai02", slug="duplicate", focus_keywords=["keyword"]),
        ]
        config = WordPressConfig("https://example.com", "u", "p")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            (folder / "bai01_bg.jpg").write_bytes(b"x")
            (folder / "bai01_1.jpg").write_bytes(b"x")
            (folder / "bai01_2.jpg").write_bytes(b"x")
            (folder / "bai02_bg.jpg").write_bytes(b"x")
            (folder / "bai02_1.jpg").write_bytes(b"x")
            results, orphans = publish_posts(
                posts,
                config,
                PosterOptions(),
                image_folder=folder,
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success", "success"])
        self.assertEqual(orphans, [])
        self.assertEqual(len(fake.created), 1)
        self.assertEqual(len(fake.updated), 1)
        self.assertEqual(fake.updated[0][0], 99)
        self.assertIn("<img", fake.updated[0][2])
        self.assertIsNotNone(fake.updated[0][3])
        self.assertIn("<img", fake.created[0][1])
        self.assertIsNotNone(fake.created[0][2])
        created_content = fake.created[0][1]
        self.assertEqual(created_content.count("<img"), 3)
        self.assertLess(created_content.index("bai01_bg.jpg"), created_content.index("bai01_1.jpg"))
        self.assertLess(created_content.index("bai01_1.jpg"), created_content.index("bai01_2.jpg"))

    def test_publish_from_excel_auto_detects_adjacent_image_zip(self) -> None:
        try:
            import pandas as pd
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeClient()
        config = WordPressConfig("https://example.com", "u", "p")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            excel_path = folder / "posts.xlsx"
            pd.DataFrame(
                [
                    {
                        "ma_bai": "bai01",
                        "Tiêu đề SEO": "First",
                        "Nội dung HTML thuần": "<p>A</p><p>B</p>",
                        "Slug": "first",
                        "Mô tả Meta SEO": "Meta",
                        "Từ khóa chính": "keyword",
                    }
                ]
            ).to_excel(excel_path, sheet_name="Bài SEO HTML", index=False)

            zip_path = folder / "VER2 (2).zip"
            with ZipFile(zip_path, "w") as archive:
                archive.writestr("VER2/bai01_bg.jpg", b"x")
                archive.writestr("VER2/bai01_1.jpg", b"x")

            results, orphans = publish_from_excel(
                excel_path,
                None,
                config,
                PosterOptions(),
                client_factory=lambda _: fake,
            )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(orphans, [])
        self.assertEqual(len(fake.created), 1)
        self.assertIn("<img", fake.created[0][1])
        self.assertIsNotNone(fake.created[0][2])

    def test_updates_existing_post_image_layout_without_republishing_fields(self) -> None:
        class ImageLayoutClient:
            def __init__(self):
                self.updated_content: tuple[int, str] | None = None

            def find_post_by_slug(self, slug: str):
                if slug.strip("/") == "first":
                    return {"id": 7, "link": "https://example.com/first"}
                return None

            def get_post(self, post_id: int):
                return {
                    "id": post_id,
                    "link": "https://example.com/first",
                    "content": {
                        "raw": '<p>A</p><p><img src="https://example.com/a.jpg" class="wp-image-7 aligncenter" width="300" /></p>'
                    },
                }

            def update_post_content(self, post_id: int, content: str):
                self.updated_content = (post_id, content)
                return {"id": post_id, "link": "https://example.com/first"}

        fake = ImageLayoutClient()

        results = update_existing_posts_image_layout(
            [Post(2, "First", "<p>Excel content is not used here</p>", slug="/first/")],
            WordPressConfig("https://example.com", "u", "p"),
            PosterOptions(image_alignment="alignleft", image_display_size="large"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertIsNotNone(fake.updated_content)
        self.assertEqual(fake.updated_content[0], 7)
        self.assertIn("alignleft", fake.updated_content[1])
        self.assertIn('width="900"', fake.updated_content[1])
        self.assertNotIn("aligncenter", fake.updated_content[1])

    def test_publish_from_excel_force_status_publish(self) -> None:
        try:
            import pandas as pd
        except ImportError:
            self.skipTest("pandas is not installed")

        fake = FakeClient()

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            excel_path = folder / "posts.xlsx"
            pd.DataFrame(
                [
                    {
                        "ma_bai": "bai01",
                        "Tiêu đề SEO": "First",
                        "Nội dung HTML thuần": "<p>A</p>",
                        "Slug": "first",
                        "status": "draft",
                    }
                ]
            ).to_excel(excel_path, sheet_name="Bài SEO HTML", index=False)

            publish_from_excel(
                excel_path,
                None,
                WordPressConfig("https://example.com", "u", "p"),
                PosterOptions(default_status="publish", force_status="publish"),
                client_factory=lambda _: fake,
            )

        self.assertEqual(fake.created[0][0].status, "publish")

    def test_bulk_publish_existing_posts_updates_status_only(self) -> None:
        class BulkClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def find_post_by_slug(self, slug: str):
                if slug.strip("/") == "first":
                    return {"id": 7, "link": "https://example.com/first"}
                return None

            def update_post_status(self, post_id: int, status: str = "publish"):
                self.updated.append((post_id, status))
                return {"id": post_id, "link": "https://example.com/first"}

        fake = BulkClient()
        results = publish_existing_posts_bulk(
            [Post(2, "First", "<p>A</p>", slug="/first/")],
            WordPressConfig("https://example.com", "u", "p"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated, [(7, "publish")])

    def test_list_website_posts_uses_client_list_posts(self) -> None:
        class WebsiteListClient:
            def list_posts(self):
                return [
                    {"id": 9, "title": {"rendered": "Bài web"}, "status": "draft"},
                ]

        messages: list[str] = []
        posts = list_website_posts(
            WordPressConfig("https://example.com", "u", "p"),
            progress_callback=messages.append,
            client_factory=lambda _: WebsiteListClient(),
        )

        self.assertEqual(posts[0]["id"], 9)
        self.assertIn("Đã tải 1 bài viết", messages[-1])

    def test_publish_website_posts_bulk_updates_selected_ids(self) -> None:
        class WebsitePublishClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def update_post_status(self, post_id: int, status: str = "publish"):
                self.updated.append((post_id, status))
                return {"id": post_id, "link": f"https://example.com/{post_id}"}

        fake = WebsitePublishClient()
        results = publish_website_posts_bulk(
            [{"id": 15, "title": {"rendered": "Bài nháp"}, "link": "https://example.com/draft"}],
            WordPressConfig("https://example.com", "u", "p"),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated, [(15, "publish")])
        self.assertEqual(results[0].row_number, 15)

    def test_update_website_posts_image_layout_updates_content_by_id(self) -> None:
        class WebsiteImageClient:
            def __init__(self):
                self.updated: list[tuple[int, str]] = []

            def get_post(self, post_id: int):
                return {
                    "id": post_id,
                    "title": {"rendered": "Bài có ảnh"},
                    "link": "https://example.com/with-image",
                    "content": {"raw": '<p><img src="https://example.com/a.jpg" /></p>'},
                }

            def update_post_content(self, post_id: int, content: str):
                self.updated.append((post_id, content))
                return {"id": post_id, "link": "https://example.com/with-image"}

        fake = WebsiteImageClient()
        results = update_website_posts_image_layout(
            [{"id": 21, "title": {"rendered": "Bài có ảnh"}}],
            WordPressConfig("https://example.com", "u", "p"),
            PosterOptions(image_alignment="aligncenter", image_display_size="medium", image_custom_width=600),
            client_factory=lambda _: fake,
        )

        self.assertEqual([result.status for result in results], ["success"])
        self.assertEqual(fake.updated[0][0], 21)
        self.assertIn("aligncenter", fake.updated[0][1])
        self.assertIn("width=\"600\"", fake.updated[0][1])

    def test_export_links_to_source_excel_appends_adjacent_link_column(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            source_path = folder / "posts.xlsx"
            output_path = folder / "posts_with_links.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Bài SEO HTML"
            worksheet.append(["Tiêu đề SEO", "Nội dung HTML thuần", "Slug"])
            worksheet.append(["Bài 1", "<p>A</p>", "bai-1"])
            worksheet.append(["Bài 2", "<p>B</p>", "bai-2"])
            workbook.save(source_path)

            exported = export_links_to_source_excel(
                source_path,
                [
                    PostResult(2, "Bài 1", "success", link="https://example.com/bai-1"),
                    PostResult(3, "Bài 2", "failed", error="Lỗi đăng"),
                ],
                output_path,
            )

            exported_workbook = load_workbook(exported)
            exported_sheet = exported_workbook["Bài SEO HTML"]

            self.assertEqual(exported, output_path)
            self.assertEqual(exported_sheet.cell(row=1, column=4).value, "Link bài viết")
            self.assertEqual(exported_sheet.cell(row=2, column=4).value, "https://example.com/bai-1")
            self.assertEqual(exported_sheet.cell(row=2, column=4).hyperlink.target, "https://example.com/bai-1")
            self.assertIsNone(exported_sheet.cell(row=3, column=4).value)

            original_workbook = load_workbook(source_path)
            self.assertEqual(original_workbook["Bài SEO HTML"].max_column, 3)

    def test_export_links_to_source_excel_reuses_existing_link_column(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")

        with tempfile.TemporaryDirectory() as directory:
            folder = Path(directory)
            source_path = folder / "posts.xlsx"
            output_path = folder / "posts_with_links.xlsx"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Bài SEO HTML"
            worksheet.append(["Tiêu đề SEO", "Nội dung HTML thuần", "Link bài viết"])
            worksheet.append(["Bài 1", "<p>A</p>", ""])
            workbook.save(source_path)

            export_links_to_source_excel(
                source_path,
                [PostResult(2, "Bài 1", "success", link="https://example.com/bai-1")],
                output_path,
            )

            exported_sheet = load_workbook(output_path)["Bài SEO HTML"]
            self.assertEqual(exported_sheet.max_column, 3)
            self.assertEqual(exported_sheet.cell(row=2, column=3).value, "https://example.com/bai-1")


if __name__ == "__main__":
    unittest.main()
