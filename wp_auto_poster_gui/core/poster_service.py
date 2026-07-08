from __future__ import annotations

from copy import copy
from datetime import datetime
import tempfile
from pathlib import Path
import time
from threading import Event
from typing import Any, Callable, Iterable
from zipfile import ZipFile

from .content_composer import compose_content_with_images, rewrite_existing_image_layout
from .excel_reader import REQUIRED_COLUMNS, _build_column_map, _normalize_label, read_posts_from_excel
from .image_matcher import SUPPORTED_IMAGE_EXTENSIONS, match_images_for_posts
from .models import Post, PostResult, PosterOptions, ProgressCallback, UploadedMedia, WordPressConfig


ClientFactory = Callable[[WordPressConfig], object]


def _default_progress(message: str) -> None:
    del message


def _make_client(config: WordPressConfig):
    from .wp_client import WordPressClient

    return WordPressClient(config)


def publish_posts(
    posts: Iterable[Post],
    config: WordPressConfig,
    options: PosterOptions,
    image_folder: str | Path | None = None,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> tuple[list[PostResult], list[Path]]:
    """Publish posts through WordPress. Tests can inject a fake client_factory."""

    progress = progress_callback or _default_progress
    post_list = list(posts)
    ma_bai_values = [post.ma_bai or "" for post in post_list if post.ma_bai]
    image_matches, orphan_files = match_images_for_posts(
        image_folder,
        ma_bai_values,
        max_images_per_post=options.max_images_per_post,
    )
    client = client_factory(config)
    results: list[PostResult] = []

    for post in post_list:
        if stop_event and stop_event.is_set():
            progress("Stopped by user")
            break

        progress(f"Processing row {post.row_number}: {post.title}")
        if options.default_status and not post.status:
            post.status = options.default_status

        try:
            duplicate_post = _find_duplicate_post(client, post.title) if options.skip_duplicates else None

            matched = image_matches.get(post.ma_bai or "", None)
            featured_media_id: int | None = None
            featured_media: UploadedMedia | None = None
            if matched and matched.background:
                media = getattr(client, "upload_media_from_path")(matched.background)
                featured_media_id = media.media_id
                featured_media = media
                progress(f"Uploaded featured image: {matched.background.name} -> {media.source_url}")
            elif post.featured_image_url:
                media = getattr(client, "upload_media_from_url")(post.featured_image_url)
                featured_media_id = media.media_id
                featured_media = media
                progress(f"Uploaded featured image URL: {media.source_url}")

            uploaded_content: list[UploadedMedia] = []
            if matched:
                for image_path in matched.content_images:
                    media = getattr(client, "upload_media_from_path")(image_path)
                    uploaded_content.append(media)
                    progress(f"Uploaded content image: {image_path.name} -> {media.source_url}")

            content = compose_content_with_images(
                post.content,
                post.title,
                uploaded_content,
                leading_images=[featured_media] if featured_media else None,
                alignment=options.image_alignment,
                display_size=options.image_display_size,
                custom_width=options.image_custom_width,
            )

            if options.skip_duplicates:
                if duplicate_post:
                    if options.dry_run:
                        results.append(PostResult(post.row_number, post.title, "success", link="dry-run-update"))
                        progress(f"Would update duplicate post: {post.title}")
                        continue
                    duplicate_id = duplicate_post.get("id")
                    if duplicate_id is not None and hasattr(client, "update_post"):
                        payload = getattr(client, "update_post")(int(duplicate_id), post, content, featured_media_id)
                        link = _post_link(payload) or _post_link(duplicate_post)
                        results.append(
                            PostResult(
                                post.row_number,
                                post.title,
                                "success",
                                link=link,
                                error="Updated existing duplicate post content and SEO fields",
                            )
                        )
                        progress(f"Updated duplicate post content, images and SEO fields: {post.title}")
                        continue
                    if duplicate_id is not None and hasattr(client, "update_post_seo"):
                        payload = getattr(client, "update_post_seo")(int(duplicate_id), post)
                        link = _post_link(payload) or _post_link(duplicate_post)
                        results.append(
                            PostResult(
                                post.row_number,
                                post.title,
                                "success",
                                link=link,
                                error="Updated existing duplicate SEO fields",
                            )
                        )
                        progress(f"Updated duplicate SEO fields: {post.title}")
                        continue
                    results.append(PostResult(post.row_number, post.title, "skipped", error="Duplicate title"))
                    progress(f"Skipped duplicate: {post.title}")
                    continue

            if options.dry_run:
                results.append(PostResult(post.row_number, post.title, "success", link="dry-run"))
                continue

            payload = getattr(client, "create_post")(post, content, featured_media_id)
            link = payload.get("link") if isinstance(payload, dict) else None
            results.append(PostResult(post.row_number, post.title, "success", link=link))
            progress(f"Posted: {post.title}")

            if config.delay_seconds > 0:
                time.sleep(config.delay_seconds)
        except Exception as exc:
            results.append(PostResult(post.row_number, post.title, "failed", error=str(exc)))
            progress(f"Failed row {post.row_number}: {exc}")

    return results, orphan_files


def _find_duplicate_post(client: object, title: str) -> dict | None:
    if hasattr(client, "find_post_by_title"):
        return getattr(client, "find_post_by_title")(title)
    if getattr(client, "check_duplicate_by_title")(title):
        return {"id": None}
    return None


def _post_link(payload: object) -> str | None:
    if not isinstance(payload, dict):
        return None
    link = payload.get("link")
    if isinstance(link, str):
        return link
    return None


def list_website_posts(
    config: WordPressConfig,
    progress_callback: ProgressCallback | None = None,
    client_factory: ClientFactory = _make_client,
) -> list[dict[str, Any]]:
    progress = progress_callback or _default_progress
    client = client_factory(config)
    if not hasattr(client, "list_posts"):
        raise RuntimeError("WordPress client không hỗ trợ tải danh sách bài viết")

    progress("Đang tải danh sách bài viết từ website...")
    posts = getattr(client, "list_posts")()
    progress(f"Đã tải {len(posts)} bài viết từ website")
    return list(posts)


def publish_website_posts_bulk(
    post_payloads: Iterable[dict[str, Any]],
    config: WordPressConfig,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> list[PostResult]:
    progress = progress_callback or _default_progress
    client = client_factory(config)
    results: list[PostResult] = []

    for index, payload in enumerate(post_payloads, start=1):
        if stop_event and stop_event.is_set():
            progress("Stopped by user")
            break

        post_id = _website_post_id(payload)
        title = _website_post_title(payload)
        progress(f"Đang xuất bản bài #{post_id or index}: {title}")
        try:
            if post_id is None:
                results.append(PostResult(index, title, "failed", error="Bài viết không có ID"))
                continue
            response_payload = getattr(client, "update_post_status")(post_id, "publish")
            link = _post_link(response_payload) or _post_link(payload)
            results.append(PostResult(post_id, title, "success", link=link))
            progress(f"Đã xuất bản: {title}")
        except Exception as exc:
            results.append(PostResult(post_id or index, title, "failed", error=str(exc)))
            progress(f"Lỗi xuất bản bài #{post_id or index}: {exc}")

    return results


def update_website_posts_image_layout(
    post_payloads: Iterable[dict[str, Any]],
    config: WordPressConfig,
    options: PosterOptions,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> list[PostResult]:
    progress = progress_callback or _default_progress
    client = client_factory(config)
    results: list[PostResult] = []

    for index, payload in enumerate(post_payloads, start=1):
        if stop_event and stop_event.is_set():
            progress("Stopped by user")
            break

        post_id = _website_post_id(payload)
        title = _website_post_title(payload)
        progress(f"Đang sửa kích thước/căn ảnh bài #{post_id or index}: {title}")
        try:
            if post_id is None:
                results.append(PostResult(index, title, "failed", error="Bài viết không có ID"))
                continue

            current_payload = getattr(client, "get_post")(post_id) if hasattr(client, "get_post") else payload
            current_content = _post_content(current_payload) or _post_content(payload)
            if not current_content:
                results.append(PostResult(post_id, title, "skipped", error="Bài không có nội dung để sửa ảnh"))
                progress(f"Bỏ qua bài không có nội dung: {title}")
                continue

            updated_content = rewrite_existing_image_layout(
                current_content,
                alignment=options.image_alignment,
                display_size=options.image_display_size,
                custom_width=options.image_custom_width,
            )
            if updated_content == current_content:
                results.append(
                    PostResult(
                        post_id,
                        title,
                        "success",
                        link=_post_link(current_payload) or _post_link(payload),
                        error="Không có ảnh cần sửa",
                    )
                )
                progress(f"Không có ảnh cần sửa: {title}")
                continue

            response_payload = getattr(client, "update_post_content")(post_id, updated_content)
            link = _post_link(response_payload) or _post_link(current_payload) or _post_link(payload)
            results.append(PostResult(post_id, title, "success", link=link))
            progress(f"Đã sửa kích thước/căn ảnh: {title}")
        except Exception as exc:
            results.append(PostResult(post_id or index, title, "failed", error=str(exc)))
            progress(f"Lỗi sửa ảnh bài #{post_id or index}: {exc}")

    return results


def _website_post_id(payload: object) -> int | None:
    if not isinstance(payload, dict):
        return None
    value = payload.get("id")
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _website_post_title(payload: object) -> str:
    if not isinstance(payload, dict):
        return "Không rõ tiêu đề"
    title = payload.get("title")
    value = ""
    if isinstance(title, dict):
        raw = title.get("raw") or title.get("rendered")
        value = str(raw or "")
    elif isinstance(title, str):
        value = title
    return _strip_html_text(value).strip() or "Không rõ tiêu đề"


def _strip_html_text(value: str) -> str:
    import html
    import re

    return html.unescape(re.sub(r"<[^>]+>", "", value))


def update_existing_posts_image_layout(
    posts: Iterable[Post],
    config: WordPressConfig,
    options: PosterOptions,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> list[PostResult]:
    progress = progress_callback or _default_progress
    client = client_factory(config)
    results: list[PostResult] = []

    for post in posts:
        if stop_event and stop_event.is_set():
            progress("Stopped by user")
            break

        progress(f"Đang sửa ảnh bài đã đăng dòng {post.row_number}: {post.title}")
        try:
            existing_post = _find_existing_post(client, post)
            if not existing_post:
                results.append(PostResult(post.row_number, post.title, "skipped", error="Không tìm thấy bài đã đăng"))
                progress(f"Không tìm thấy bài đã đăng: {post.title}")
                continue

            post_id = existing_post.get("id")
            if post_id is None:
                results.append(PostResult(post.row_number, post.title, "failed", error="Bài đã đăng không có ID"))
                continue

            current_payload = getattr(client, "get_post")(int(post_id)) if hasattr(client, "get_post") else existing_post
            current_content = _post_content(current_payload) or _post_content(existing_post)
            if not current_content:
                results.append(PostResult(post.row_number, post.title, "skipped", error="Bài không có nội dung để sửa ảnh"))
                progress(f"Bỏ qua bài không có nội dung: {post.title}")
                continue

            updated_content = rewrite_existing_image_layout(
                current_content,
                alignment=options.image_alignment,
                display_size=options.image_display_size,
                custom_width=options.image_custom_width,
            )
            if updated_content == current_content:
                results.append(PostResult(post.row_number, post.title, "success", link=_post_link(current_payload), error="Không có ảnh cần sửa"))
                progress(f"Không có ảnh cần sửa: {post.title}")
                continue

            if hasattr(client, "update_post_content"):
                payload = getattr(client, "update_post_content")(int(post_id), updated_content)
            else:
                payload = getattr(client, "update_post")(int(post_id), post, updated_content, None)
            link = _post_link(payload) or _post_link(current_payload) or _post_link(existing_post)
            results.append(PostResult(post.row_number, post.title, "success", link=link))
            progress(f"Đã sửa kích thước/căn ảnh: {post.title}")
        except Exception as exc:
            results.append(PostResult(post.row_number, post.title, "failed", error=str(exc)))
            progress(f"Lỗi sửa ảnh dòng {post.row_number}: {exc}")

    return results


def _find_existing_post(client: object, post: Post) -> dict | None:
    if post.slug and hasattr(client, "find_post_by_slug"):
        found = getattr(client, "find_post_by_slug")(post.slug)
        if found:
            return found
    return _find_duplicate_post(client, post.title)


def _post_content(payload: object) -> str | None:
    if not isinstance(payload, dict):
        return None
    content = payload.get("content")
    if isinstance(content, str):
        return content
    if isinstance(content, dict):
        for key in ("raw", "rendered"):
            value = content.get(key)
            if isinstance(value, str) and value.strip():
                return value
    return None


def publish_from_excel(
    excel_path: str | Path,
    image_folder: str | Path | None,
    config: WordPressConfig,
    options: PosterOptions,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> tuple[list[PostResult], list[Path]]:
    posts = read_posts_from_excel(excel_path, default_status=options.default_status)
    if options.force_status:
        for post in posts:
            post.status = options.force_status
    progress = progress_callback or _default_progress
    prepared_source, cleanup = _prepare_image_source(excel_path, image_folder, posts, options, progress)
    try:
        return publish_posts(
            posts,
            config,
            options,
            image_folder=prepared_source,
            progress_callback=progress_callback,
            stop_event=stop_event,
            client_factory=client_factory,
        )
    finally:
        if cleanup is not None:
            cleanup.cleanup()


def publish_existing_posts_bulk(
    posts: Iterable[Post],
    config: WordPressConfig,
    progress_callback: ProgressCallback | None = None,
    stop_event: Event | None = None,
    client_factory: ClientFactory = _make_client,
) -> list[PostResult]:
    progress = progress_callback or _default_progress
    client = client_factory(config)
    results: list[PostResult] = []

    for post in posts:
        if stop_event and stop_event.is_set():
            progress("Stopped by user")
            break

        progress(f"Đang xuất bản bài dòng {post.row_number}: {post.title}")
        try:
            existing_post = _find_existing_post(client, post)
            if not existing_post:
                results.append(PostResult(post.row_number, post.title, "skipped", error="Không tìm thấy bài đã đăng"))
                progress(f"Không tìm thấy bài đã đăng: {post.title}")
                continue

            post_id = existing_post.get("id")
            if post_id is None:
                results.append(PostResult(post.row_number, post.title, "failed", error="Bài đã đăng không có ID"))
                continue

            payload = getattr(client, "update_post_status")(int(post_id), "publish")
            link = _post_link(payload) or _post_link(existing_post)
            results.append(PostResult(post.row_number, post.title, "success", link=link))
            progress(f"Đã xuất bản: {post.title}")
        except Exception as exc:
            results.append(PostResult(post.row_number, post.title, "failed", error=str(exc)))
            progress(f"Lỗi xuất bản dòng {post.row_number}: {exc}")

    return results


def _prepare_image_source(
    excel_path: str | Path,
    image_source: str | Path | None,
    posts: list[Post],
    options: PosterOptions,
    progress: ProgressCallback,
) -> tuple[Path | None, tempfile.TemporaryDirectory | None]:
    ma_bai_values = [post.ma_bai or "" for post in posts if post.ma_bai]
    if not ma_bai_values:
        return Path(image_source) if image_source else None, None

    selected = Path(image_source) if image_source else None
    if selected:
        if selected.suffix.lower() == ".zip" and selected.exists():
            temp_dir = tempfile.TemporaryDirectory(prefix="wp-auto-poster-images-")
            extracted = _extract_matching_images_from_zip(selected, ma_bai_values, Path(temp_dir.name))
            progress(f"Đã đọc file ZIP ảnh: {selected.name} ({extracted} ảnh khớp)")
            return Path(temp_dir.name), temp_dir
        try:
            matches, _orphans = match_images_for_posts(selected, ma_bai_values, options.max_images_per_post)
            if _matched_image_count(matches) > 0:
                return selected, None
            progress(f"Thư mục ảnh đang chọn không khớp mã bài: {selected}")
        except FileNotFoundError:
            progress(f"Không tìm thấy thư mục/file ảnh: {selected}")

    detected = _find_matching_image_source(Path(excel_path).parent, ma_bai_values)
    if not detected:
        return selected, None

    if detected.suffix.lower() == ".zip":
        temp_dir = tempfile.TemporaryDirectory(prefix="wp-auto-poster-images-")
        extracted = _extract_matching_images_from_zip(detected, ma_bai_values, Path(temp_dir.name))
        progress(f"Tự nhận file ZIP ảnh: {detected.name} ({extracted} ảnh khớp)")
        return Path(temp_dir.name), temp_dir

    progress(f"Tự nhận thư mục ảnh: {detected}")
    return detected, None


def prepare_image_source(
    excel_path: str | Path,
    image_source: str | Path | None,
    posts: list[Post],
    options: PosterOptions,
    progress: ProgressCallback,
) -> tuple[Path | None, tempfile.TemporaryDirectory | None]:
    return _prepare_image_source(excel_path, image_source, posts, options, progress)


def _find_matching_image_source(base_dir: Path, ma_bai_values: list[str]) -> Path | None:
    if not base_dir.exists():
        return None
    candidates = [
        item
        for item in base_dir.iterdir()
        if item.is_dir() or item.suffix.lower() == ".zip"
    ]
    candidates.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    for candidate in candidates:
        try:
            if candidate.suffix.lower() == ".zip":
                if _zip_match_count(candidate, ma_bai_values) > 0:
                    return candidate
            else:
                matches, _orphans = match_images_for_posts(candidate, ma_bai_values)
                if _matched_image_count(matches) > 0:
                    return candidate
        except Exception:
            continue
    return None


def _zip_match_count(zip_path: Path, ma_bai_values: list[str]) -> int:
    count = 0
    prefixes = tuple(f"{code}_".lower() for code in ma_bai_values if code)
    with ZipFile(zip_path) as archive:
        for entry in archive.infolist():
            name = Path(entry.filename).name
            suffix = Path(name).suffix.lower()
            if not name or suffix not in SUPPORTED_IMAGE_EXTENSIONS:
                continue
            if name.lower().startswith(prefixes):
                count += 1
    return count


def _extract_matching_images_from_zip(zip_path: Path, ma_bai_values: list[str], output_dir: Path) -> int:
    output_dir.mkdir(parents=True, exist_ok=True)
    prefixes = tuple(f"{code}_".lower() for code in ma_bai_values if code)
    count = 0
    with ZipFile(zip_path) as archive:
        for entry in archive.infolist():
            name = Path(entry.filename).name
            suffix = Path(name).suffix.lower()
            if not name or suffix not in SUPPORTED_IMAGE_EXTENSIONS:
                continue
            if not name.lower().startswith(prefixes):
                continue
            target = output_dir / name
            with archive.open(entry) as source, target.open("wb") as destination:
                destination.write(source.read())
            count += 1
    return count


def _matched_image_count(matches: dict) -> int:
    total = 0
    for match in matches.values():
        if match.background:
            total += 1
        total += len(match.content_images)
    return total


def export_results_to_excel(results: list[PostResult], output_path: str | Path) -> None:
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("pandas and openpyxl are required to export Excel reports") from exc

    frame = pd.DataFrame(
        [
            {
                "row": result.row_number,
                "title": result.title,
                "status": result.status,
                "link": result.link,
                "error": result.error,
            }
            for result in results
        ]
    )
    frame.to_excel(output_path, index=False)


def export_links_to_source_excel(
    source_path: str | Path,
    results: list[PostResult],
    output_path: str | Path | None = None,
) -> Path:
    """Copy the original workbook and append post links next to the source data."""

    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on environment
        raise RuntimeError("openpyxl is required to export links into the source Excel file") from exc

    source = Path(source_path)
    if not source.exists():
        raise FileNotFoundError(f"Excel file does not exist: {source}")

    output = Path(output_path) if output_path else _default_link_export_path(source)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source)
    worksheet = _find_post_worksheet(workbook)
    link_column = _link_column(worksheet)
    header_cell = worksheet.cell(row=1, column=link_column)
    header_cell.value = "Link bài viết"

    if link_column > 1:
        previous_header = worksheet.cell(row=1, column=link_column - 1)
        if previous_header.has_style:
            header_cell._style = copy(previous_header._style)
        header_cell.font = copy(previous_header.font)
        header_cell.fill = copy(previous_header.fill)
        header_cell.border = copy(previous_header.border)
        header_cell.alignment = copy(previous_header.alignment)
        header_cell.number_format = previous_header.number_format
        previous_letter = get_column_letter(link_column - 1)
        current_letter = get_column_letter(link_column)
        worksheet.column_dimensions[current_letter].width = max(
            worksheet.column_dimensions[previous_letter].width or 12,
            28,
        )

    for result in results:
        cell = worksheet.cell(row=result.row_number, column=link_column)
        cell.value = result.link or ""
        if result.link:
            cell.hyperlink = result.link
            cell.style = "Hyperlink"

    workbook.save(output)
    return output


def _default_link_export_path(source: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return source.with_name(f"{source.stem}_ket_qua_dang_bai_{timestamp}{source.suffix}")


def _find_post_worksheet(workbook):
    worksheets = list(workbook.worksheets)
    preferred = [
        worksheet for worksheet in worksheets if "bai seo html" in _normalize_label(worksheet.title)
    ]
    for worksheet in [*preferred, *[sheet for sheet in worksheets if sheet not in preferred]]:
        headers = [cell.value for cell in worksheet[1]]
        column_map = _build_column_map(headers)
        if REQUIRED_COLUMNS <= set(column_map):
            return worksheet
    if worksheets:
        return worksheets[0]
    raise RuntimeError("Workbook does not contain any worksheet")


def _link_column(worksheet) -> int:
    rightmost_header = 0
    for cell in worksheet[1]:
        if cell.value not in (None, ""):
            rightmost_header = cell.column
            if _normalize_label(cell.value) == "link bai viet":
                return cell.column
    return rightmost_header + 1 if rightmost_header else 1
